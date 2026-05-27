"""
evidence_checker.py

完整的证据检查流水线：
1. EvidenceLoader —— 从 supporting Excel 按 sheet 提取嵌入图片
2. OCRExtractor —— RapidOCR 文本检测（含 bbox 坐标）
3. EvidenceChecker —— 表头锚定 + 全量兜底双保险比对
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zipfile import ZipFile

import numpy as np
import pandas as pd
from PIL import Image
from rapidocr_onnxruntime import RapidOCR


# ============================================================
# 0. Sheet → Node 映射
# ============================================================

# 当前 YUMCZOO 的 supporting Excel 中已有的 sheet
# Key = 检查节点名, Value = supporting.xlsx 中的 sheet 名
SHEET_TO_NODE: Dict[str, str] = {
    "应用层管理员账号": "需审阅应用层管理员账号",
    "数据库账号": "需审阅数据库账号",
    "发版工具发版账号": "需审阅发版工具发版账号",
}

# 反向映射：Node → Sheet
NODE_TO_SHEET: Dict[str, str] = {v: k for k, v in SHEET_TO_NODE.items()}

# 完整模板（所有节点 → 对应的 sheet 名）
# 后续可改为从 configs/supporting_sheet_mapping.xlsx 读取
FULL_NODE_TO_SHEET: Dict[str, str] = {
    "需审阅应用层管理员账号": "应用层管理员账号",
    "需上传应用层账号截图": "应用层管理员账号",
    "需审阅数据库账号": "数据库账号",
    "需审阅发版工具发版账号": "发版工具发版账号",
    "需审阅发版工具后台账号": "发版工具后台账号",
    "需审阅应用层服务器账号": "应用层服务器账号",
    "需审阅数据库服务器账号": "数据库服务器账号",
    "需上传开发人员清单": "开发人员清单",
    "开发人员在生产环境无只读以上权限": "开发人员权限检查",
    "无前后台管理员SOD问题": "SOD检查",
}

# 不同节点对应的表头锚点关键词
ANCHOR_KEYWORDS: Dict[str, Dict[str, List[str]]] = {
    "需审阅应用层管理员账号": {
        "account_column": ["User/group", "用户/组", "用户名", "账号", "Account", "User"],
        "role_column": ["角色", "Role", "权限", "Permission"],
    },
    "需审阅数据库账号": {
        "account_column": [
            "User", "user", "username", "Username",
            "用户名", "账号", "Account", "account",
        ],
        "host_column": ["Host", "host", "主机", "地址"],
    },
    "需审阅发版工具发版账号": {
        "account_column": ["User", "用户名", "用户", "账号", "Account", "Name"],
        "role_column": ["角色", "Role", "权限"],
    },
    "需审阅发版工具后台账号": {
        "account_column": ["User", "用户名", "账号", "Account"],
        "role_column": ["角色", "Role", "权限"],
    },
    "需审阅应用层服务器账号": {
        "account_column": ["User", "用户名", "账号", "Account", "账号名"],
        "owner_column": ["使用人", "Owner", "用户"],
    },
    "需审阅数据库服务器账号": {
        "account_column": ["User", "用户名", "账号", "Account", "账号名"],
        "owner_column": ["使用人", "Owner", "用户"],
    },
}


# ============================================================
# 1. 文本工具
# ============================================================

def clean_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).replace("\u3000", " ").strip()
    text = re.sub(r"[ \t]+", " ", text)
    return text


def normalize_account(value: str) -> str:
    """账号弱标准化：去空格、特殊字符、统一大写。"""
    text = clean_text(value).upper()
    # 基础去噪
    text = re.sub(r"[\s_\-/\\（）()【】\[\].:：]+", "", text)
    return text


def normalize_account_mysql(value: str) -> str:
    """
    MySQL 终端 OCR 特化归一化。

    RapidOCR 对终端字体的常见误读：
    - % → * 或 10 或 1O
    - @ → Q 或丢失
    - | → I 或 1 或丢失
    """
    text = clean_text(value).upper()
    # 基础去噪
    text = re.sub(r"[\s_\-/\\（）()【】\[\].:：]+", "", text)
    # MySQL 特化修复
    text = text.replace("*", "%")       # * → %
    text = text.replace("Q", "@")       # Q → @ (在 user@host 上下文中)
    text = re.sub(r"^[|1I]+", "", text)  # 去掉开头的表格线残留
    text = re.sub(r"[|1I]+$", "", text)  # 去掉结尾的表格线残留
    # 修复 "1O" → "10" → 这种误读少见，先保留
    return text


def fuzzy_normalize(value: str) -> str:
    """
    极度弱化归一化：用于全局兜底。
    去掉几乎所有符号，只留字母数字。
    """
    text = clean_text(value).upper()
    text = re.sub(r"[^A-Z0-9]", "", text)
    return text


def fuzzy_char_ratio(a: str, b: str) -> float:
    """
    字符重叠率：短字符串中有多少字母数字出现在长字符串中。

    比 Jaccard 更适合 OCR 场景——OCR 多读/漏读几个字符不应惩罚匹配。

    例如 REPL@% vs rep10* (OCR):
    - 去特殊符号后: REPL vs REP10
    - 短串 REPL 的 4 个字符全部在 REP10 中出现 → 1.0
    """
    a_chars = re.sub(r"[^A-Z0-9]", "", a.upper())
    b_chars = re.sub(r"[^A-Z0-9]", "", b.upper())

    if not a_chars or not b_chars:
        return 0.0

    # 取较短的一方做覆盖检查
    if len(a_chars) <= len(b_chars):
        shorter, longer = a_chars, b_chars
    else:
        shorter, longer = b_chars, a_chars

    # 短串中每个字符在长串中出现的比例
    hits = sum(1 for ch in shorter if ch in longer)
    return hits / len(shorter)


# ============================================================
# 2. EvidenceLoader
# ============================================================

@dataclass
class ImageBatch:
    """从 supporting Excel 中提取的一批截图。"""
    node: str
    sheet_name: str
    images: List[Image.Image]
    image_count: int
    source_file: str


class EvidenceLoader:
    """从 supporting Excel 的指定 sheet 中提取嵌入图片。"""

    def __init__(self, node_to_sheet: Optional[Dict[str, str]] = None):
        self.node_to_sheet = node_to_sheet or FULL_NODE_TO_SHEET

    def load_for_node(self, supporting_path: str | Path, node: str) -> Optional[ImageBatch]:
        """
        根据 node 找到对应 sheet，提取该 sheet 中所有嵌入图片。
        返回 ImageBatch，如果该 node 在当前 supporting 中没有对应 sheet 则返回 None。
        """
        sheet_name = self.node_to_sheet.get(node, "")
        if not sheet_name:
            return None

        images = self._extract_images_from_sheet(supporting_path, sheet_name)

        if not images:
            return None

        return ImageBatch(
            node=node,
            sheet_name=sheet_name,
            images=images,
            image_count=len(images),
            source_file=str(supporting_path),
        )

    def load_all(self, supporting_path: str | Path) -> Dict[str, ImageBatch]:
        """加载所有 node 对应的图片批次。"""
        result: Dict[str, ImageBatch] = {}
        for node in self.node_to_sheet:
            batch = self.load_for_node(supporting_path, node)
            if batch is not None:
                result[node] = batch
        return result

    def _extract_images_from_sheet(self, filepath: str | Path, sheet_name: str) -> List[Image.Image]:
        """
        从 Excel 指定 sheet 提取嵌入图片。

        实现方法：
        1. 用 openpyxl 读取，通过 sheet._images 获取图片
        2. 从 zipfile 读取原始图片数据
        3. 通过 drawing rels 映射 sheet → drawing → image

        但由于 openpyxl 的 _images 在不同版本中不稳定，
        这里用 zipfile + drawing rels 做精确映射。
        """
        filepath = Path(filepath)
        if not filepath.exists():
            return []

        sheet_index = self._find_sheet_index(filepath, sheet_name)
        if sheet_index is None:
            return []

        try:
            with ZipFile(filepath, 'r') as z:
                # 读取该 sheet 的 drawing 引用
                drawing_path = self._read_sheet_drawing(z, sheet_index)
                if drawing_path is None:
                    return []

                # 从 drawing rels 获取图片引用
                image_refs = self._read_drawing_image_refs(z, drawing_path)

                # 提取图片
                images: List[Image.Image] = []
                for img_path in image_refs:
                    if img_path in z.namelist():
                        data = z.read(img_path)
                        img = Image.open(io.BytesIO(data))
                        if img.mode == 'RGBA':
                            img = img.convert('RGB')
                        images.append(img)

                return images
        except Exception:
            return []

    def _find_sheet_index(self, filepath: Path, sheet_name: str) -> Optional[int]:
        """找到 sheet 在 workbook 中的索引（1-based）。"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            for i, name in enumerate(wb.sheetnames, start=1):
                if name == sheet_name:
                    wb.close()
                    return i
            wb.close()
            return None
        except Exception:
            return None

    def _read_sheet_drawing(self, z: ZipFile, sheet_index: int) -> Optional[str]:
        """读取 sheetX.xml.rels 中的 drawing 路径。"""
        rels_path = f"xl/worksheets/_rels/sheet{sheet_index}.xml.rels"
        if rels_path not in z.namelist():
            return None

        content = z.read(rels_path).decode('utf-8', errors='replace')

        # 找 drawing 的 Target
        match = re.search(
            r'Type="[^"]*drawing"[^>]*Target="([^"]+)"',
            content
        )
        if match:
            target = match.group(1)
            # 转成 zip 内的绝对路径
            if target.startswith('../'):
                target = 'xl/' + target[3:]
            elif not target.startswith('xl/'):
                target = 'xl/drawings/' + target
            return target

        return None

    def _read_drawing_image_refs(self, z: ZipFile, drawing_path: str) -> List[str]:
        """从 drawing rels 中读取图片引用路径。"""
        # drawing1.xml → drawing1.xml.rels
        rels_dir = '/'.join(drawing_path.split('/')[:-1])
        rels_file = drawing_path.split('/')[-1]
        rels_path = f"{rels_dir}/_rels/{rels_file}.rels"

        if rels_path not in z.namelist():
            return []

        content = z.read(rels_path).decode('utf-8', errors='replace')

        # 找所有 image Target
        targets = re.findall(
            r'Type="[^"]*image"[^>]*Target="([^"]+)"',
            content
        )

        images: List[str] = []
        rels_dir_parts = rels_dir.split('/')  # e.g. ["xl", "drawings"]

        for target in targets:
            # 规范化路径：消除 ../
            if target.startswith('../'):
                # 从 drawing 目录出发解析相对路径
                target_parts = target.split('/')
                base_parts = list(rels_dir_parts)
                for part in target_parts:
                    if part == '..':
                        base_parts.pop()
                    else:
                        base_parts.append(part)
                resolved = '/'.join(base_parts)
            elif target.startswith('/'):
                resolved = target.lstrip('/')
            else:
                resolved = f"xl/media/{target}"

            if resolved in z.namelist():
                images.append(resolved)

        return images


# ============================================================
# 3. OCRExtractor
# ============================================================

@dataclass
class TextBox:
    """OCR 检测到的一个文本块。"""
    text: str
    bbox: Tuple[float, float, float, float]  # (x1, y1, x2, y2)
    confidence: float

    @property
    def x_center(self) -> float:
        return (self.bbox[0] + self.bbox[2]) / 2

    @property
    def y_center(self) -> float:
        return (self.bbox[1] + self.bbox[3]) / 2

    @property
    def x_min(self) -> float:
        return self.bbox[0]

    @property
    def y_min(self) -> float:
        return self.bbox[1]

    @property
    def y_max(self) -> float:
        return self.bbox[3]


@dataclass
class OCRPage:
    """单张图片的 OCR 结果。"""
    text_blocks: List[TextBox]
    full_text: str
    image_index: int  # 在 batch 中的序号


class OCRExtractor:
    """RapidOCR 封装。"""

    def __init__(self):
        self.engine = RapidOCR()

    def extract(self, batch: ImageBatch) -> List[OCRPage]:
        """对一批图片逐张运行 OCR。"""
        pages: List[OCRPage] = []
        for i, img in enumerate(batch.images):
            page = self._ocr_one(img, i)
            pages.append(page)
        return pages

    def _ocr_one(self, img: Image.Image, index: int) -> OCRPage:
        result, _ = self.engine(img)

        blocks: List[TextBox] = []
        if result is None:
            return OCRPage(text_blocks=[], full_text="", image_index=index)

        for item in result:
            # RapidOCR 返回格式: [bbox四点, text, confidence]
            # bbox 格式: [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
            poly = item[0]
            text = str(item[1])
            confidence = float(item[2])

            # 四点转矩形 (x1,y1,x2,y2)
            xs = [p[0] for p in poly]
            ys = [p[1] for p in poly]
            bbox = (min(xs), min(ys), max(xs), max(ys))

            blocks.append(TextBox(text=text, bbox=bbox, confidence=confidence))

        # 构建全量文本池
        full_text = "\n".join(b.text for b in blocks)

        return OCRPage(text_blocks=blocks, full_text=full_text, image_index=index)


# ============================================================
# 4. EvidenceChecker
# ============================================================

@dataclass
class AccountMatch:
    """单个 expected account 的比对结果。"""
    expected_account: str
    matched_account: str = ""
    matched: bool = False
    match_method: str = ""  # "spatial_anchor" | "global_exact" | "global_fuzzy" | "none"
    confidence: float = 0.0
    source_image: int = -1  # 在哪张图片中找到的


@dataclass
class NodeCheckResult:
    """一个节点的完整检查结果。"""
    task_id: str
    node: str
    system: str
    expected_accounts: List[str]
    actual_accounts: List[str]
    passed: List[AccountMatch] = field(default_factory=list)
    spatial_mismatch: List[AccountMatch] = field(default_factory=list)
    missing: List[AccountMatch] = field(default_factory=list)
    extra: List[str] = field(default_factory=list)

    @property
    def overall_status(self) -> str:
        if self.missing and not self.passed:
            return "FAIL"
        if self.missing:
            return "REVIEW"
        if self.spatial_mismatch:
            return "REVIEW_SPATIAL"
        return "PASS"


class EvidenceChecker:
    """双保险比对引擎：表头锚定 + 全量兜底。"""

    def __init__(self, anchor_keywords: Optional[Dict[str, Dict[str, List[str]]]] = None):
        self.anchor_keywords = anchor_keywords or ANCHOR_KEYWORDS

    def check(
        self,
        task_id: str,
        node: str,
        system: str,
        expected_accounts: List[str],
        ocr_pages: List[OCRPage],
    ) -> NodeCheckResult:
        """
        执行比对：
        1. 第一道：在每个 OCR page 上做表头锚定，提取列内账号
        2. 第二道：未匹配的 expected account 做全量文本池兜底
        """
        result = NodeCheckResult(
            task_id=task_id,
            node=node,
            system=system,
            expected_accounts=expected_accounts,
            actual_accounts=[],
        )

        all_anchor_texts: List[str] = []  # 锚定方式提取到的所有 actual 账号
        all_global_text: str = ""  # 所有页面的全量文本池

        for page in ocr_pages:
            # 第一道：表头锚定
            anchor_texts = self._spatial_anchor_extract(
                page, node
            )
            all_anchor_texts.extend(anchor_texts)

            # 收集全量文本
            all_global_text += page.full_text + "\n"

        # 标准化锚定文本
        # 数据库节点使用 MySQL 特化归一化（修复 %→* 等 OCR 误读）
        is_db_node = "数据库账号" in node
        norm_func = normalize_account_mysql if is_db_node else normalize_account
        anchor_normalized = [norm_func(t) for t in all_anchor_texts]
        result.actual_accounts = sorted(set(anchor_normalized))

        # 对每个 expected account 做匹配
        for expected in expected_accounts:
            exp_norm = norm_func(expected)

            # 第一道：在锚定结果中找
            match = self._match_in_list(exp_norm, anchor_normalized, all_anchor_texts)
            if match:
                match.match_method = "spatial_anchor"
                match.confidence = 1.0 if exp_norm == norm_func(match.matched_account) else 0.85
                result.passed.append(match)
                continue

            # 第二道：全量文本池兜底（含 MySQL 纠错）
            match = self._global_fuzzy_match(expected, exp_norm, all_global_text, is_db_node=is_db_node)
            if match:
                match.match_method = "global_fuzzy"
                result.spatial_mismatch.append(match)
                continue

            # 彻底找不到
            result.missing.append(AccountMatch(
                expected_account=expected,
                match_method="none",
                confidence=0.0,
            ))

        # 找出 OCR 中多出来的（不在 expected 中的）
        exp_norms = {norm_func(e) for e in expected_accounts}
        for actual in result.actual_accounts:
            if actual not in exp_norms:
                result.extra.append(actual)

        return result

    def _spatial_anchor_extract(self, page: OCRPage, node: str) -> List[str]:
        """
        表头空间锚定提取。

        策略：
        1. 找到表头锚点词（如 "User/group"、"账号"）
        2. 以锚点为中心，推断该列的 X 区间
        3. 提取锚点下方、列区间内的所有文本
        4. 同行合并
        """
        anchor_config = self.anchor_keywords.get(node, {})
        account_keywords = anchor_config.get("account_column", ["User", "账号", "用户名", "Account"])

        blocks = page.text_blocks
        if not blocks:
            return []

        # Step 1: 找到所有候选锚点
        anchor_blocks: List[TextBox] = []
        for block in blocks:
            for kw in account_keywords:
                if kw.lower() in block.text.lower():
                    anchor_blocks.append(block)
                    break

        if not anchor_blocks:
            return []

        # Step 2: 对每个锚点，推断列范围并提取
        extracted: List[str] = []
        for anchor in anchor_blocks:
            col_zone = self._compute_column_zone(blocks, anchor)
            col_texts = self._extract_column_texts(blocks, col_zone, anchor.y_max)
            merged = self._merge_same_row(col_texts)
            extracted.extend(merged)

        return extracted

    def _compute_column_zone(
        self,
        blocks: List[TextBox],
        anchor: TextBox,
        tolerance_ratio: float = 0.5,
    ) -> Tuple[float, float]:
        """
        根据锚点和其他文本块推断该列的 X 边界。

        策略：
        1. 以锚点 X 中心为起点
        2. 向左右扫描，找列间隙（连续 X 范围无文本）
        3. 如果找不到清晰边界，使用锚点 ± 50% 锚点宽度
        """
        anchor_width = anchor.bbox[2] - anchor.bbox[0]
        default_margin = anchor_width * tolerance_ratio

        # 收集所有文本块的 X 范围
        all_x_ranges = sorted([
            (b.bbox[0], b.bbox[2])
            for b in blocks
            if b.bbox[1] > anchor.bbox[1] - 20  # 只看锚点附近的
        ], key=lambda r: r[0])

        if not all_x_ranges:
            return (anchor.x_min - default_margin, anchor.bbox[2] + default_margin)

        # 找左边界：锚点 X 中心左侧最近的文本右边界
        left_boundary = 0.0
        for x1, x2 in all_x_ranges:
            if x2 < anchor.x_center and x2 > left_boundary:
                left_boundary = x2

        # 找右边界：锚点 X 中心右侧最近的文本左边界
        right_boundary = float('inf')
        for x1, x2 in all_x_ranges:
            if x1 > anchor.x_center and x1 < right_boundary:
                right_boundary = x1

        # 如果边界太远，用默认边距
        col_left = max(left_boundary, anchor.x_min - default_margin)
        col_right = min(right_boundary, anchor.bbox[2] + default_margin)

        return (col_left, col_right)

    def _extract_column_texts(
        self,
        blocks: List[TextBox],
        col_zone: Tuple[float, float],
        anchor_y_max: float,
    ) -> List[TextBox]:
        """提取锚点下方、列区间内的所有文本块，按 Y 坐标排序。"""
        col_left, col_right = col_zone

        candidates = [
            b for b in blocks
            if b.y_min >= anchor_y_max - 5  # 锚点下方（小容差）
            and col_left - 15 <= b.x_center <= col_right + 15  # ±15px 水平容错
        ]

        candidates.sort(key=lambda b: b.y_center)
        return candidates

    def _merge_same_row(
        self,
        texts: List[TextBox],
        y_threshold: float = 10.0,
    ) -> List[str]:
        """同行文本合并：Y 坐标差 < threshold 的视为同一行。"""
        if not texts:
            return []

        merged: List[str] = []
        current_row: List[TextBox] = [texts[0]]
        current_y = texts[0].y_center

        for tb in texts[1:]:
            if abs(tb.y_center - current_y) < y_threshold:
                current_row.append(tb)
                # 用加权平均更新当前行 Y
                current_y = (current_y * len(current_row) + tb.y_center) / (len(current_row) + 1)
            else:
                # 输出上一行
                row_text = " ".join(t.text for t in current_row)
                merged.append(row_text)
                current_row = [tb]
                current_y = tb.y_center

        # 最后一行
        if current_row:
            row_text = " ".join(t.text for t in current_row)
            merged.append(row_text)

        return merged

    def _match_in_list(
        self,
        exp_norm: str,
        anchor_norms: List[str],
        anchor_originals: List[str],
    ) -> Optional[AccountMatch]:
        """在锚定结果列表中查找匹配。"""
        for i, norm in enumerate(anchor_norms):
            # 精确匹配
            if exp_norm == norm:
                return AccountMatch(
                    expected_account="",
                    matched_account=anchor_originals[i],
                    matched=True,
                )
            # 子串包含
            if exp_norm in norm or norm in exp_norm:
                return AccountMatch(
                    expected_account="",
                    matched_account=anchor_originals[i],
                    matched=True,
                )
        return None

    def _global_fuzzy_match(
        self,
        expected: str,
        exp_norm: str,
        all_text: str,
        is_db_node: bool = False,
    ) -> Optional[AccountMatch]:
        """
        全量文本池弱化匹配兜底。

        步骤：
        1. 标准化后精确匹配
        2. 子串包含（expected in global or global in expected）
        3. 对数据库节点：尝试 MySQL 特化归一化再匹配
        4. 短账号（≤3 字符）必须通过 \b 边界检查
        """
        global_norm = normalize_account(all_text)

        # 对数据库节点，额外生成 MySQL 纠错版全局文本
        global_mysql: Optional[str] = None
        if is_db_node:
            global_mysql = normalize_account_mysql(all_text)

        # 精确匹配
        if exp_norm in all_text.upper() or exp_norm in global_norm:
            return AccountMatch(
                expected_account=expected,
                matched_account=expected,
                matched=True,
            )

        # MySQL 纠错精确匹配
        if global_mysql and exp_norm in global_mysql:
            return AccountMatch(
                expected_account=expected,
                matched_account=expected,
                matched=True,
            )

        # 子串包含：逐行检查
        lines = all_text.split('\n')
        for line in lines:
            line_norm = normalize_account(line)
            if not line_norm:
                continue

            # expected 在 line 中（标准归一化）
            if exp_norm in line_norm:
                if len(exp_norm) <= 3:
                    if not self._short_account_defense(exp_norm, line_norm):
                        continue
                return AccountMatch(
                    expected_account=expected,
                    matched_account=line.strip(),
                    matched=True,
                )

            # MySQL 纠错后匹配
            if is_db_node:
                line_mysql = normalize_account_mysql(line)
                if line_mysql and exp_norm in line_mysql:
                    return AccountMatch(
                        expected_account=expected,
                        matched_account=line.strip(),
                        matched=True,
                    )
                # 反向：MySQL 归一化后的行在 expected 中
                if len(exp_norm) > 4 and line_mysql in exp_norm:
                    return AccountMatch(
                        expected_account=expected,
                        matched_account=line.strip(),
                        matched=True,
                    )

            # line 在 expected 中（标准归一化，expected 必须有一定长度）
            if len(exp_norm) > 4 and line_norm in exp_norm:
                return AccountMatch(
                    expected_account=expected,
                    matched_account=line.strip(),
                    matched=True,
                )

        # 第三道：字符级模糊匹配
        # 条件：短字符串 ≥ 3 个字母数字，且 ≥ 80% 的字符在对方中出现
        for line in lines:
            line_norm = normalize_account(line)
            if not line_norm or len(line_norm) < 3:
                continue
            sim = fuzzy_char_ratio(exp_norm, line_norm)
            if sim >= 0.80:
                match = AccountMatch(
                    expected_account=expected,
                    matched_account=line.strip(),
                    matched=True,
                )
                match.confidence = sim * 0.55  # 模糊匹配置信度打折
                return match

        return None

    def _short_account_defense(self, short_account: str, line_norm: str) -> bool:
        """
        短账号（≤3 字符）防御。

        规则：
        1. 用 \b 单词边界正则
        2. 匹配文本长度不应远超账号长度（ratio ≤ 2x）
        """
        try:
            pattern = re.compile(r'\b' + re.escape(short_account) + r'\b')
            if not pattern.search(line_norm):
                return False
        except re.error:
            return False

        # 长度比率检查
        if len(line_norm) > len(short_account) * 2:
            return False

        return True


# ============================================================
# 5. 主流程：对 YUMCZOO 执行完整检查
# ============================================================

def run_evidence_check(
    supporting_path: str | Path,
    execution_plan_path: str | Path,
    output_path: Optional[str | Path] = None,
    system_filter: Optional[str] = None,
) -> Dict[str, NodeCheckResult]:
    """
    主入口：
    1. 读取 execution_plan.xlsx 中的 expected_accounts 和 evidence_plan
    2. 加载 supporting Excel 中的截图
    3. OCR 扫描
    4. 双保险比对
    5. 写回 compare_result

    system_filter: 可选，只处理指定系统的节点（如 "YUMC Zoo"）。
    """
    supporting_path = Path(supporting_path)
    execution_plan_path = Path(execution_plan_path)

    if output_path is None:
        output_path = execution_plan_path

    # 读取 execution plan
    evidence_df = pd.read_excel(execution_plan_path, sheet_name="evidence_plan")
    expected_df = pd.read_excel(execution_plan_path, sheet_name="expected_accounts")
    compare_df = pd.read_excel(execution_plan_path, sheet_name="compare_result")

    # 初始化各模块
    loader = EvidenceLoader()
    ocr = OCRExtractor()
    checker = EvidenceChecker()

    all_results: Dict[str, NodeCheckResult] = {}

    # 按节点聚合 expected accounts
    node_expected: Dict[str, List[str]] = {}
    for _, row in expected_df.iterrows():
        if row.get("ExpectedStatus", "") == "KEEP":
            task_id = clean_text(row.get("TaskID", ""))
            account = clean_text(row.get("AccountStd", ""))
            if task_id and account:
                if task_id not in node_expected:
                    node_expected[task_id] = []
                if account not in node_expected[task_id]:
                    node_expected[task_id].append(account)

    # 逐节点处理
    for _, plan_row in evidence_df.iterrows():
        task_id = clean_text(plan_row.get("TaskID", ""))
        node = clean_text(plan_row.get("Node", ""))
        system = clean_text(plan_row.get("System", ""))

        # 按系统过滤
        if system_filter and system_filter.lower() not in system.lower():
            continue

        # 检查该节点是否已有 supporting 截图
        sheet_name = FULL_NODE_TO_SHEET.get(node, "")
        if not sheet_name:
            continue

        # 尝试加载截图
        batch = loader.load_for_node(supporting_path, node)

        # 「需上传应用层账号截图」是证据存在性检查，不跑 OCR
        # 只要能加载到截图（sheet 存在且有嵌入图片）→ PASS
        if node == "需上传应用层账号截图":
            result = NodeCheckResult(
                task_id=task_id, node=node, system=system,
                expected_accounts=[], actual_accounts=[],
            )
            if batch is not None and batch.image_count > 0:
                result.passed.append(AccountMatch(
                    expected_account="", matched_account="",
                    matched=True, match_method="evidence_exists",
                    confidence=1.0,
                ))
            else:
                result.missing.append(AccountMatch(
                    expected_account="截图未上传",
                    match_method="none", confidence=0.0,
                ))
            all_results[task_id] = result
            print(f"[{node}] sheet={sheet_name} | "
                  f"images={batch.image_count if batch else 0} | "
                  f"status={result.overall_status}")
            continue

        if batch is None:
            continue

        # OCR
        pages = ocr.extract(batch)

        # 获取 expected accounts
        expected = node_expected.get(task_id, [])

        # 比对
        result = checker.check(task_id, node, system, expected, pages)
        all_results[task_id] = result

        print(f"[{node}] expected={len(expected)} | "
              f"ocr_pages={len(pages)} | "
              f"passed={len(result.passed)} | "
              f"spatial_mismatch={len(result.spatial_mismatch)} | "
              f"missing={len(result.missing)} | "
              f"extra={len(result.extra)} | "
              f"status={result.overall_status}")

    # 写回 compare_result
    _write_compare_results(compare_df, all_results, execution_plan_path, output_path)

    return all_results


def _write_compare_results(
    compare_df: pd.DataFrame,
    results: Dict[str, NodeCheckResult],
    execution_plan_path: Path,
    output_path: Path,
) -> None:
    """将检查结果写回 compare_result sheet。"""
    result_map: Dict[str, NodeCheckResult] = results

    new_rows: List[Dict[str, Any]] = []
    for _, row in compare_df.iterrows():
        task_id = clean_text(row.get("TaskID", ""))
        row_dict = row.to_dict()

        if task_id in result_map:
            r = result_map[task_id]
            row_dict["ActualAccounts"] = " | ".join(r.actual_accounts)
            row_dict["ActualCount"] = len(r.actual_accounts)
            row_dict["MissingAccounts"] = " | ".join(m.expected_account for m in r.missing)
            row_dict["ExtraAccounts"] = " | ".join(r.extra)
            row_dict["CheckResult"] = r.overall_status
            row_dict["Remark"] = _build_remark(r)
        else:
            # 没有 supporting 截图的节点保持原状
            pass

        new_rows.append(row_dict)

    new_df = pd.DataFrame(new_rows)

    # 写回（保留其他 sheet 不变）
    with pd.ExcelWriter(
        output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        new_df.to_excel(writer, sheet_name="compare_result", index=False)


def _build_remark(r: NodeCheckResult) -> str:
    parts: List[str] = []
    if r.passed:
        accounts = [m.matched_account for m in r.passed if m.matched_account]
        parts.append(f"锚定匹配({len(r.passed)}): {', '.join(accounts[:5])}")
    if r.spatial_mismatch:
        accounts = [m.expected_account for m in r.spatial_mismatch]
        parts.append(f"全量兜底({len(r.spatial_mismatch)}): {', '.join(accounts[:5])}")
    if r.missing:
        accounts = [m.expected_account for m in r.missing]
        parts.append(f"缺失({len(r.missing)}): {', '.join(accounts[:5])}")
    if r.extra:
        parts.append(f"多余({len(r.extra)}): {', '.join(r.extra[:5])}")
    return " | ".join(parts)


# ============================================================
# 6. CLI
# ============================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Run evidence check on supporting screenshots.")
    parser.add_argument(
        "--supporting",
        required=True,
        help="supporting Excel 文件路径（如 supporting/YumcZoo/支持性附件截图.xlsx）",
    )
    parser.add_argument(
        "--execution-plan",
        default="output/execution_plan.xlsx",
        help="execution_plan.xlsx 路径",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出路径（默认覆写原 execution_plan）",
    )
    parser.add_argument(
        "--system",
        default=None,
        help="只处理指定系统（如 'YUMC Zoo'）",
    )

    args = parser.parse_args()

    results = run_evidence_check(
        supporting_path=args.supporting,
        execution_plan_path=args.execution_plan,
        output_path=args.output,
        system_filter=args.system,
    )

    print(f"\nDone. {len(results)} nodes checked.")
